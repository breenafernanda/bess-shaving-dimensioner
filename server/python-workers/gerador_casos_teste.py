"""
MÓDULO: Gerador de Casos de Teste

Gera arquivos Excel realistas com curvas de carga industriais
para simular diferentes cenários de empresas.

Uso:
    python gerador_casos_teste.py --stage 3 --severity moderado --days 30
"""

import random
import json
from datetime import datetime, timedelta
from typing import Tuple, List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ============================================================================
# CONFIGURAÇÕES
# ============================================================================

# Ranges de demanda contratada por estágio de empresa (em kW)
COMPANY_STAGES = {
    1: {"name": "Pequeno Comércio", "min_kw": 10, "max_kw": 50},
    2: {"name": "Pequena Indústria", "min_kw": 50, "max_kw": 150},
    3: {"name": "Indústria Média", "min_kw": 150, "max_kw": 500},
    4: {"name": "Indústria Grande", "min_kw": 500, "max_kw": 1500},
    5: {"name": "Indústria Pesada", "min_kw": 1500, "max_kw": 5000},
}

# Variabilidade por nível de severidade
SEVERITY_LEVELS = {
    "leve": {"min_var": -0.05, "max_var": 0.05, "peak_factor": 1.10},
    "moderado": {"min_var": -0.15, "max_var": 0.15, "peak_factor": 1.25},
    "grave": {"min_var": -0.30, "max_var": 0.30, "peak_factor": 1.50},
}

# Prefixos para nomes de empresas
COMPANY_PREFIXES = [
    "Metalúrgica",
    "Indústria",
    "Fábrica",
    "Manufatura",
    "Processadora",
    "Usinagem",
    "Plásticos",
    "Têxtil",
    "Alimentos",
    "Bebidas",
]

# Nomes para empresas
COMPANY_NAMES = [
    "Silva",
    "Santos",
    "Oliveira",
    "Costa",
    "Ferreira",
    "Gomes",
    "Martins",
    "Pereira",
    "Souza",
    "Rocha",
    "Alves",
    "Ribeiro",
    "Carvalho",
    "Barbosa",
    "Monteiro",
]

# Sufixos para empresas
COMPANY_SUFFIXES = [
    "LTDA",
    "S.A.",
    "Ind.",
    "Manufatureira",
    "Processadora",
    "Usinagem",
    "Indústria",
]


# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

def gerar_nome_empresa() -> str:
    """
    Gera um nome de empresa aleatório realista.
    
    Returns:
        str: Nome da empresa (ex: "Metalúrgica Silva LTDA")
    """
    prefixo = random.choice(COMPANY_PREFIXES)
    nome = random.choice(COMPANY_NAMES)
    sufixo = random.choice(COMPANY_SUFFIXES)
    
    return f"{prefixo} {nome} {sufixo}"


def selecionar_demanda_contratada(stage: int) -> float:
    """
    Seleciona uma demanda contratada aleatória dentro do range do estágio.
    
    Args:
        stage: Estágio da empresa (1-5)
        
    Returns:
        float: Demanda contratada em kW
    """
    if stage not in COMPANY_STAGES:
        raise ValueError(f"Estágio inválido: {stage}. Deve ser 1-5.")
    
    config = COMPANY_STAGES[stage]
    demanda = random.uniform(config["min_kw"], config["max_kw"])
    
    # Arredondar para valor mais realista
    return round(demanda, 1)


def gerar_curva_base(hora: int) -> float:
    """
    Gera a curva de carga base (normalizada 0-1) para cada hora do dia.
    
    Padrão industrial típico:
    - 00-06: Consumo mínimo (madrugada)
    - 06-12: Subida gradual (manhã)
    - 12-13: Pequena queda (almoço)
    - 13-18: Recuperação (tarde)
    - 18-22: Pico máximo (ponta)
    - 22-24: Queda (noite)
    
    Args:
        hora: Hora do dia (0-23)
        
    Returns:
        float: Valor normalizado (0-1)
    """
    if 0 <= hora < 6:
        # Madrugada: consumo mínimo
        return 0.35
    
    elif 6 <= hora < 12:
        # Manhã: subida gradual
        # Sobe de 0.35 até 0.83 em 6 horas
        return 0.35 + (hora - 6) * 0.08
    
    elif 12 <= hora < 13:
        # Almoço: pequena queda
        return 0.73
    
    elif 13 <= hora < 18:
        # Tarde: recuperação
        # Sobe de 0.73 até 0.98 em 5 horas
        return 0.73 + (hora - 13) * 0.05
    
    elif 18 <= hora < 22:
        # Ponta: pico máximo
        return 0.98
    
    else:  # 22 <= hora < 24
        # Noite: queda para madrugada
        # Cai de 0.98 até 0.35 em 2 horas
        return 0.98 - (hora - 22) * 0.315


def aplicar_variabilidade(valor_base: float, severidade: str) -> float:
    """
    Aplica variabilidade ao valor base conforme o nível de severidade.
    
    Args:
        valor_base: Valor normalizado (0-1)
        severidade: Nível de severidade ('leve', 'moderado', 'grave')
        
    Returns:
        float: Valor com variabilidade aplicada
    """
    if severidade not in SEVERITY_LEVELS:
        raise ValueError(f"Severidade inválida: {severidade}")
    
    config = SEVERITY_LEVELS[severidade]
    
    # Adicionar ruído aleatório
    ruido = random.uniform(config["min_var"], config["max_var"])
    valor_variado = valor_base + ruido
    
    # Garantir que não seja negativo e não ultrapasse muito o máximo
    valor_variado = max(0.1, min(1.5, valor_variado))
    
    return valor_variado


def converter_para_kw(valor_pu: float, demanda_contratada: float) -> float:
    """
    Converte valor normalizado (pu) para kW.
    
    Args:
        valor_pu: Valor normalizado (0-1)
        demanda_contratada: Demanda contratada em kW
        
    Returns:
        float: Potência em kW
    """
    return valor_pu * demanda_contratada


def gerar_timestamps(data_inicio: datetime, dias: int) -> List[datetime]:
    """
    Gera lista de timestamps horários para o período especificado.
    
    Args:
        data_inicio: Data de início
        dias: Número de dias
        
    Returns:
        List[datetime]: Lista de timestamps
    """
    timestamps = []
    data_atual = data_inicio
    
    for _ in range(dias * 24):  # 24 horas por dia
        timestamps.append(data_atual)
        data_atual += timedelta(hours=1)
    
    return timestamps


def gerar_curva_carga(
    demanda_contratada: float,
    severidade: str,
    dias: int,
    data_inicio: datetime = None
) -> Tuple[List[datetime], List[float]]:
    """
    Gera a curva de carga completa para o período especificado.
    
    Args:
        demanda_contratada: Demanda contratada em kW
        severidade: Nível de severidade
        dias: Número de dias
        data_inicio: Data de início (padrão: hoje)
        
    Returns:
        Tuple[List[datetime], List[float]]: Timestamps e potências em kW
    """
    if data_inicio is None:
        data_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    timestamps = gerar_timestamps(data_inicio, dias)
    potencias = []
    
    for timestamp in timestamps:
        hora = timestamp.hour
        
        # Gerar curva base
        valor_base = gerar_curva_base(hora)
        
        # Aplicar variabilidade
        valor_variado = aplicar_variabilidade(valor_base, severidade)
        
        # Converter para kW
        potencia_kw = converter_para_kw(valor_variado, demanda_contratada)
        
        # Arredondar para 1 casa decimal
        potencia_kw = round(potencia_kw, 1)
        
        potencias.append(potencia_kw)
    
    return timestamps, potencias


def gerar_arquivo_excel(
    timestamps: List[datetime],
    potencias: List[float],
    nome_empresa: str,
    caminho_saida: str
) -> str:
    """
    Gera arquivo Excel no formato Elspec.
    
    Args:
        timestamps: Lista de timestamps
        potencias: Lista de potências em kW
        nome_empresa: Nome da empresa
        caminho_saida: Caminho para salvar o arquivo
        
    Returns:
        str: Caminho do arquivo criado
    """
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Adicionar headers
    ws["A1"] = "Time stamp"
    ws["B1"] = f"[kW] Active Power Total (Cycle) Average, {nome_empresa} - WYE"
    
    # Adicionar dados
    for i, (timestamp, potencia) in enumerate(zip(timestamps, potencias), start=2):
        # Formatar timestamp como DD/MM/YYYY HH:MM:SS.000000
        timestamp_str = timestamp.strftime("%d/%m/%Y %H:%M:%S.000000")
        ws[f"A{i}"] = timestamp_str
        ws[f"B{i}"] = potencia
    
    # Ajustar largura das colunas
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    
    # Salvar arquivo
    wb.save(caminho_saida)
    
    return caminho_saida


def gerar_caso_teste(
    stage: int,
    severidade: str,
    dias: int,
    data_inicio: datetime = None,
    caminho_saida: str = None
) -> Dict:
    """
    Gera um caso de teste completo (arquivo Excel + metadados).
    
    Args:
        stage: Estágio da empresa (1-5)
        severidade: Nível de severidade ('leve', 'moderado', 'grave')
        dias: Número de dias a simular
        data_inicio: Data de início (padrão: hoje)
        caminho_saida: Caminho para salvar (padrão: auto-gerado)
        
    Returns:
        Dict: Metadados do caso gerado
    """
    # Validar entrada
    if stage not in COMPANY_STAGES:
        raise ValueError(f"Estágio inválido: {stage}. Deve ser 1-5.")
    
    if severidade not in SEVERITY_LEVELS:
        raise ValueError(f"Severidade inválida: {severidade}")
    
    if dias < 1 or dias > 365:
        raise ValueError(f"Dias inválido: {dias}. Deve ser 1-365.")
    
    # Definir data de início
    if data_inicio is None:
        data_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Gerar nome da empresa
    nome_empresa = gerar_nome_empresa()
    
    # Selecionar demanda contratada
    demanda_contratada = selecionar_demanda_contratada(stage)
    
    # Gerar curva de carga
    timestamps, potencias = gerar_curva_carga(
        demanda_contratada,
        severidade,
        dias,
        data_inicio
    )
    
    # Definir caminho de saída
    if caminho_saida is None:
        nome_arquivo = f"caso_teste_{stage}_{severidade}_{dias}dias.xlsx"
        caminho_saida = f"/uploads/{nome_arquivo}"
    
    # Gerar arquivo Excel
    caminho_arquivo = gerar_arquivo_excel(
        timestamps,
        potencias,
        nome_empresa,
        caminho_saida
    )
    
    # Calcular estatísticas
    potencia_max = max(potencias)
    potencia_min = min(potencias)
    potencia_media = sum(potencias) / len(potencias)
    
    # Retornar metadados
    return {
        "sucesso": True,
        "arquivo": caminho_arquivo,
        "nome_empresa": nome_empresa,
        "stage": stage,
        "stage_descricao": COMPANY_STAGES[stage]["name"],
        "severidade": severidade,
        "dias": dias,
        "data_inicio": data_inicio.isoformat(),
        "data_fim": (data_inicio + timedelta(days=dias - 1)).isoformat(),
        "demanda_contratada_kw": demanda_contratada,
        "potencia_maxima_kw": round(potencia_max, 1),
        "potencia_minima_kw": round(potencia_min, 1),
        "potencia_media_kw": round(potencia_media, 1),
        "total_pontos": len(potencias),
    }


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Gerador de casos de teste para BESS Peak Shaving"
    )
    parser.add_argument(
        "--stage",
        type=int,
        default=3,
        choices=[1, 2, 3, 4, 5],
        help="Estágio da empresa (1-5)"
    )
    parser.add_argument(
        "--severity",
        type=str,
        default="moderado",
        choices=["leve", "moderado", "grave"],
        help="Nível de severidade"
    )
    parser.add_argument(
        "--days",
        type=int,
        default=30,
        help="Número de dias a simular (1-365)"
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Caminho de saída do arquivo"
    )
    
    args = parser.parse_args()
    
    # Gerar caso de teste
    resultado = gerar_caso_teste(
        stage=args.stage,
        severidade=args.severity,
        dias=args.days,
        caminho_saida=args.output
    )
    
    # Exibir resultado
    print(json.dumps(resultado, indent=2, ensure_ascii=False))
